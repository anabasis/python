#!/usr/bin/env python

# pip install virtualenv
# virtualenv exceltrans
# \exceltrans\Scripts\activate
# python -m pip install --upgrade pip
# pip install xlsxwriter
# pip install openpyxl
# pip install pandas

# Openpyxl <https://doitnow-man.tistory.com/159>
# Pandas <https://www.delftstack.com/ko/howto/python-pandas/how-to-add-new-column-to-existing-dataframe-in-python-pandas/>
# curl "https://openapi.naver.com/v1/papago/n2mt" -H "Content-Type: application/x-www-form-urlencoded; charset=UTF-8" -H "X-Naver-Client-Id: msl8RsSf0ro6hhXFT8cR" -H "X-Naver-Client-Secret: ESoXCEZlHI" -d "source=en&target=ko&text=Advanced Threat Detection Security Monitoring Compliance" -v
# curl -v --get 'https://dapi.kakao.com/v2/translation/translate' -d 'src_lang=en' -d 'target_lang=kr' --data-urlencode 'query=Advanced Threat Detection Security Monitoring Compliance' -H 'Authorization: KakaoAK 53436d8595e369d48bcf14aa59cf15f6'

from openpyxl import load_workbook
from openpyxl import Workbook
from itertools import islice
import numpy as np
import pandas as pd
import os
import sys
import json
import time
import urllib.request

def xlsx_read(filename):

    wb = load_workbook(filename)
    sheet_names = wb.sheetnames
    print(sheet_names)

    for sheet_name in sheet_names:
        print(sheet_name)

        sheet_xlsx = wb[    sheet_name]

        print(sheet_xlsx.max_row)
        print(sheet_xlsx.max_column)

        for row in range(1, sheet_xlsx.max_row):
            for col in range(1, sheet_xlsx.max_column):
                print(sheet_xlsx.cell(row,col).value)

    wb.close()
    return print(sheet_names)

def xlsx_sheet_read(filename, sheetname='Sheet1'):
    wb = load_workbook(filename)
    sheet_xlsx = wb[sheetname]

    print(sheet_xlsx.max_row)
    print(sheet_xlsx.max_column)

    for row in range(1, sheet_xlsx.max_row):
        for col in range(1, sheet_xlsx.max_column):
            print(sheet_xlsx.cell(row,col).value)

    wb.close()
    return print(sheetname)

def xlsx_sheet_trans(filename, sheetname='Sheet1', cellrange=[]):

    wb = load_workbook(filename)
    sheet_xlsx = wb[sheetname]

    print('## SIZE  ##' + '#'*18)
    print(sheet_xlsx.max_row)
    print(sheet_xlsx.max_column)
    print('#'*25)
    total_nf = np.empty((0))

    for field in cellrange :
        data = sheet_xlsx[field] # tuple
        if len(total_nf) == 0 :
            total_nf = np.array(data)
        else :
            total_nf = np.hstack((total_nf, np.array(data)))
    #print(total_nf.shape)

    fields = [ [ col.value for col in row] for row in total_nf[:1]]
    contents = [ [ col.value for col in row] for row in total_nf[1:]]
    df = pd.DataFrame(contents, columns=fields)

    ws = wb.create_sheet()
    ws.title = sheetname + '_한글'
    dict = {}

    for k in range(len(total_nf)) :
        ws.cell(row=k+1, column=1).value = total_nf[k,0].value

    for i in range(len(total_nf)) :
        for j in range(len(total_nf[0])) :
            if total_nf[i,j].value in dict :
                ws.cell(row=i+1, column=j+2).value = dict[total_nf[i,j].value]
            else:
                ws.cell(row=i+1, column=j+2).value = kakao_trans(total_nf[i,j].value)
                #ws.cell(row=i+1, column=j+2).value = "한글번역"
                dict[total_nf[i,j].value] = total_nf[i,j].value
                time.sleep(2)

    wb.save(filename)
    wb.close()
    return df

def naver_trans(sentence) :

    # curl "https://openapi.naver.com/v1/papago/n2mt" -H "Content-Type: application/x-www-form-urlencoded; charset=UTF-8" -H "X-Naver-Client-Id: msl8RsSf0ro6hhXFT8cR" -H "X-Naver-Client-Secret: ESoXCEZlHI" -d "source=en&target=ko&text=Advanced Threat Detection Security Monitoring Compliance" -v

    #client_id = "YOUR_CLIENT_ID"
    #client_secret = "YOUR_CLIENT_SECRET"
    client_id = "msl8RsSf0ro6hhXFT8cR"
    client_secret = "ESoXCEZlHI"

    encText = urllib.parse.quote(sentence)
    data = "source=ko&target=en&text=" + encText
    url = "https://openapi.naver.com/v1/language/translate"
    request = urllib.request.Request(url)
    request.add_header("X-Naver-Client-Id",client_id)
    request.add_header("X-Naver-Client-Secret",client_secret)
    response = urllib.request.urlopen(request, data=data.encode("utf-8"))
    rescode = response.getcode()
    if(rescode==200):
        response_body = response.read().decode('utf-8')
        sentence = ''
        for row  in json.loads(response_body)["translated_text"] :
            sentence = sentence + "\n" + ''.join(row)
    else:
        #print("Error Code:" + rescode)
        sentence = "Error Code:" + rescode
    return sentence
    # write stuff to database or something...

def kakao_trans(sentence):

    # curl -v --get 'https://dapi.kakao.com/v2/translation/translate' -d 'src_lang=en' -d 'target_lang=kr' --data-urlencode 'query=Advanced Threat Detection Security Monitoring Compliance' -H 'Authorization: KakaoAK 53436d8595e369d48bcf14aa59cf15f6'
    #encText = urllib.parse.quote("Advanced Threat Detection Security Monitoring Compliance")
    rest_api_key = '53436d8595e369d48bcf14aa59cf15f6'
    payload = {'src_lang':'en','target_lang':'kr','query':sentence}
    url = "https://dapi.kakao.com/v2/translation/translate"

    request = urllib.request.Request(url)
    request.add_header("Authorization",'KakaoAK '+ rest_api_key)

    response = urllib.request.urlopen(request, urllib.parse.urlencode(payload).encode('utf-8'))
    rescode = response.getcode()
    if(rescode==200):
        response_body = response.read().decode('utf-8')
        sentence = ''
        for row  in json.loads(response_body)["translated_text"] :
            sentence = sentence + "\n" + ''.join(row)
    else:
        #print("Error Code:" + rescode)
        sentence = "Error Code:" + rescode
    return sentence
    # write stuff to database or something...

if __name__ == "__main__":

    #xlsx_read('.\\data\\ES_시나리오_한글_20200910.xlsx')
    #xlsx_sheet_trans('.\\data\\ES_시나리오_한글_20200910.xlsx','ANALYTICS',['B2:B493','F2:G493','P2:P493'])
    #main_df = xlsx_sheet_trans('.\\data\\ES_시나리오_한글_20200910.xlsx','ANALYTICS',['B2:B493','F2:G493','P2:P493'])
    main_df = xlsx_sheet_trans('.\\data\\ES_시나리오_한글_20200910.xlsx','ANALYTICS',['B2:B493'])
    #print(kakao_trans(sentence))
    #print(kakao_trans(sentence))
